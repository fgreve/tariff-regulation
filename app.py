import pandas as pd
import numpy as np
import os
import json
from openpyxl import load_workbook
from unidecode import unidecode

from app_aux import nombre_ies, output_cols, output_cols_aux, label, costos_matricula


def extract_ies_all():
    xls = os.listdir("data/BBDD/IES/")
    for xl in xls:
        print(xl)
        extract_ies(xl)


def extract_ies(xl_file):
    """ extrae desde el archivo fuente de las ies a a carpeta /extract/IES/"""
    # wb = "data/BBDD/IES/PUC.xlsx"
    nombre_institucion = xl_file.split(".")[0]
    path_file_wb = os.path.join("data/BBDD/IES/", xl_file)

    data_univ = {}

    workbook = load_workbook(filename=path_file_wb, read_only=True, data_only=True)
    wss = workbook.sheetnames
    print(nombre_institucion, " | ", nombre_ies[nombre_institucion], " | ",
          "cantidad de carreras (codigos unicos): ", len(wss) - 2)

    for ws in wss:
        sheet = workbook[ws]

        if ws == "Costos Centrales":
            datos = sheet["A3:G20"]
            institucion = sheet['B1'].value
        elif ws == "Costos de matrícula":
            datos = sheet["A3:G10"]
            institucion = sheet['B1'].value
        else:
            datos = sheet["A5:G30"]
            institucion = sheet['B1'].value
            carrera = sheet['B2'].value
            jornada = sheet['B3'].value
            sede = sheet['B4'].value

        data = []
        for x in datos:
            row = []
            for y in x:
                row.append(y.value)

            data.append(row)

        if ws == "Costos Centrales":
            data_univ[ws] = {"institucion": institucion, "data": data}
            print(ws, " | ", institucion)
        elif ws == "Costos de matrícula":
            data_univ[ws] = {"institucion": institucion, "data": data}
            print(ws, " | ", institucion)
        else:
            info = {"institucion": institucion, "carrera": carrera, "jornada": jornada, "sede": sede}
            data_univ[ws] = {"info": info, "data": data}
            print(ws, " | ", institucion, " | ", carrera, " | ", jornada, " | ", sede)

    path_file_json = os.path.join("data/extract/", nombre_institucion + ".json")
    with open(path_file_json, 'w', encoding="utf8") as outfile:
        json.dump(data_univ, outfile, indent=4, ensure_ascii=False)

    workbook.close()


def get_dta_from_bbdd(file):
    """ a partir del nombre de un .dta retorna un input desde lacarpeta 0. Carpeta Comisión de Expertos\1. BB. DD  """
    path = "../0. Carpeta Comisión de Expertos/1. BB. DD/"
    file = file + ".dta"
    path_file = os.path.join(path, file)
    df = pd.read_stata(path_file)
    return df


def transform_codigo_unico_to_df(ies, codigo_unico):
    path_file = os.path.join("data/extract/", ies + ".json")

    with open(path_file, encoding="utf8") as f:
        diccionario = json.load(f)

    if (codigo_unico != 'Costos Centrales') and (codigo_unico != 'Costos de matrícula'):
        info = diccionario[codigo_unico]['info']
        # print(info)
        data = diccionario[codigo_unico]['data']
        df = pd.DataFrame(data, columns=['item', 'texto', '2015', '2016', '2017', '2018', '2019'])

        df = df.dropna(subset=['item'])
        df = df.dropna(subset=['2015', '2016', '2017', '2018', '2019'], how='all')
        df['item'] = df['item'].apply(lambda string: unidecode(string.replace(" ", "_").
                                                               replace(".", "_").replace("-", "_").lower()))
        df = df.drop(columns=['texto'])
        # df = df.dropna(axis=1, how='all')
        df.index = df.item
        df = df.drop(columns=['item'])
        df = df.transpose()
        df["codigo_unico"] = codigo_unico
        df["anio"] = df.index
        df["institucion"] = info["institucion"]
        df["carrera"] = info["carrera"]
        df["jornada"] = info["jornada"]
        df["sede"] = info["sede"]
        # print(df)

    elif codigo_unico == 'Costos Centrales':
        institucion = diccionario[codigo_unico]['institucion']
        data = diccionario[codigo_unico]['data']
        df = pd.DataFrame(data, columns=['item', 'texto', '2015', '2016', '2017', '2018', '2019'])

        df = df.dropna(subset=['item'])
        df = df.dropna(subset=['2015', '2016', '2017', '2018', '2019'], how='all')
        df['item'] = df['item'].apply(lambda string: unidecode(string.replace(" ", "_").
                                                               replace(".", "_").replace("-", "_").lower()))
        df = df.drop(columns=['texto'])
        df = df.dropna(axis=1, how='all')
        df = df[df.item != "seccion_c"]
        df.index = df.item
        df = df.drop(columns=['item'])
        df = df.transpose()
        df["anio"] = df.index
        # print(df)

    elif codigo_unico == 'Costos de matrícula':
        institucion = diccionario[codigo_unico]['institucion']
        data = diccionario[codigo_unico]['data']

        df = pd.DataFrame(data, columns=['item', 'texto', '2015', '2016', '2017', '2018', '2019'])
        df = df.dropna(subset=['item'])
        # df = df.dropna(subset=['2015', '2016', '2017', '2018', '2019'], how='all')
        df['item'] = df['item'].apply(lambda string: unidecode(string.replace(" ", "_").
                                                               replace(".", "_").replace("-", "_").lower()))

        df = df.drop(columns=['texto'])

        df.index = df.item
        df = df.drop(columns=['item'])
        df = df.transpose()
        df["anio"] = df.index

    return df


def transform_ies_to_df(ies):
    path_file = os.path.join("data/extract/", ies + ".json")

    with open(path_file, encoding="utf8") as f:
        data = json.load(f)

    codigos = data.keys()
    dfs = []
    for codigo_unico in codigos:
        if (codigo_unico != 'Costos Centrales') and (codigo_unico != 'Costos de matrícula'):
            print(ies, " | ", codigo_unico)
            df = transform_codigo_unico_to_df(ies, codigo_unico)
            dfs.append(df)

    df = pd.concat(dfs, axis='rows')
    path_file = os.path.join("data/transform/codigo_unico/", ies + ".csv")
    df.to_csv(path_file, index=False, encoding="utf-8", header=True)

    df_costos_centrales = transform_codigo_unico_to_df(ies, 'Costos Centrales')
    path_file = os.path.join("data/transform/costos_centrales/", ies + ".csv")
    df_costos_centrales.to_csv(path_file, index=False, encoding="utf-8", header=True)

    df_costos_matricula = transform_codigo_unico_to_df(ies, 'Costos de matrícula')
    path_file = os.path.join("data/transform/costos_de_matricula/", ies + ".csv")
    df_costos_matricula.to_csv(path_file, index=False, encoding="utf-8", header=True)

    # df.merge(df_costos_centrales, how='left')
    # df.merge(df_costos_matricula, how='left')


def transform_ies_to_df_all():
    for file in os.listdir("data/extract/"):
        ies = file.split(".")[0]
        print(ies)
        transform_ies_to_df(ies)


def make_sin_formato():
    output = get_output()
    output = output.drop(['seccion_4', 'seccion_3', 'seccion_2'], axis=1)
    columnas_output = list(output.columns)

    path = os.path.join("../../../Desktop/0. Carpeta Comisión de Expertos/1. BB. DD/", "output.dta")
    output_sub = pd.read_stata(path)
    # output_sub.to_csv(os.path.join("data/", "output_sub.csv"), index=False, encoding="utf-8-sig", header=True)

    inacap = pd.DataFrame(output_sub.loc[output_sub.institucion == "IP INACAP"])
    inacap["ies"] = "IP-INACAP"
    inacap = inacap[columnas_output]
    inacap.to_csv(os.path.join("data/BBDD/IES-sin-formato/", "ip-inacap.csv"), index=False, encoding="utf-8-sig",
                  header=True)

    duoc = pd.DataFrame(output_sub[output_sub.institucion == "IP DUOC UC"])
    duoc["ies"] = "IP-DUOC-UC"
    duoc = duoc[columnas_output]
    duoc.to_csv(os.path.join("data/BBDD/IES-sin-formato/", "ip-duoc-uc.csv"), index=False, encoding="utf-8-sig",
                header=True)

    cft_inacap = pd.DataFrame(output_sub[output_sub.institucion == "CFT INACAP"])
    cft_inacap["ies"] = "CFT-INACAP"
    cft_inacap = cft_inacap[columnas_output]
    cft_inacap.to_csv(os.path.join("data/BBDD/IES-sin-formato/", "cft-inacap.csv"), index=False, encoding="utf-8-sig",
                      header=True)


def make_output():
    """ este es un input para output_rep """
    path = "data/transform/"
    codigo_unico = pd.read_csv(os.path.join(path, "codigo_unico.csv"))
    costos_centrales = pd.read_csv(os.path.join(path, "costos_centrales.csv"))
    costos_de_matricula = pd.read_csv(os.path.join(path, "costos_de_matricula.csv"))

    codigo_unico.set_index(['ies', 'anio'])
    costos_centrales.set_index(['ies', 'anio'])
    costos_de_matricula.set_index(['ies', 'anio'])

    output = pd.concat([codigo_unico, costos_centrales], axis=1)
    output = pd.concat([output, costos_de_matricula], axis=1)

    output = output.loc[:, ~output.columns.duplicated()]
    output.to_csv(os.path.join("data/", "output.csv"), index=False, encoding="utf-8-sig", header=True)


def make_output_rep():
    output = get_output()

    ip_duoc_uc = pd.read_csv(os.path.join("data/BBDD/IES-sin-formato/", "ip-duoc-uc.csv"))
    ip_inacap = pd.read_csv(os.path.join("data/BBDD/IES-sin-formato/", "ip-inacap.csv"))
    cft_inacap = pd.read_csv(os.path.join("data/BBDD/IES-sin-formato/", "cft-inacap.csv"))

    output_rep = pd.concat([output, ip_duoc_uc, ip_inacap, cft_inacap], axis='rows')

    # Matrícula por carrera - Necesaria para obtener el costo per cápita de las carreras
    matricula = get_dta_from_bbdd("matricula")

    matricula.set_index(['codigo_unico', 'anio'])
    output_rep.set_index(['codigo_unico', 'anio'])
    # output_rep = pd.concat([output_rep, matricula], axis=1)
    output_rep = pd.merge(output_rep, matricula, how='left')
    output_rep = output_rep.loc[:, ~output_rep.columns.duplicated()]
    print(output_rep.head())

    # Auxiliares para datos vacios codigo unico
    # bys codigo_unico : egen max_matricula = max(aux_matricula)
    # output_rep.groupby(["codigo_unico"])["aux_matricula"].agg(['max'])

    oferta_2020 = get_dta_from_bbdd("oferta_2020")
    output_rep = pd.merge(output_rep, oferta_2020, how='left', on='codigo_unico')

    # Control por nivel de carrera
    output_rep.loc[output_rep.nivel_carrera.isin([1, 2]), 'subsistema'] = 1
    output_rep.loc[output_rep.nivel_carrera.isin([0, 3, 4]), 'subsistema'] = 2

    # Mantener carreras a regular
    output_rep = output_rep[
        output_rep.oecdsubarea.isin(["Servicios Personales", "Formación de Personal Docente", "Derecho"])]

    output_rep = output_rep.replace('-', None)

    # CFT PUCV escribio en millones
    variables = ["sub_item_c_2_1", "sub_item_c_3_1", "sub_item_c_3_2", "sub_item_c_3_3", "sub_item_c_3_4", "item_m_2",
                 "item_m_3", "item_m_4"]
    for var in variables:
        output_rep.loc[output_rep.institucion == "CFT PUCV", var] = output_rep[var] * 1000

    # PUCV, UANTO, UTALCA escribio en pesos
    variables = ["sub_item_c_2_1", "sub_item_c_2_2", "sub_item_c_3_1", "sub_item_c_3_2", "sub_item_c_3_3",
                 "sub_item_c_3_4", "item_m_2", "item_m_3", "item_m_4", "sub_item_1_1_1", "sub_item_1_1_2",
                 "sub_item_1_1_3", "item_1_2", "item_2_1", "item_2_2", "item_2_3", "item_2_4", "item_2_5", "item_3_1",
                 "item_3_2", "item_3_3", "item_3_4", "item_3_5", "item_3_6", "item_4_1", "item_4_2", "sub_item_c_1_2"]
    for var in variables:
        output_rep[var] = output_rep[var].astype(float)
        output_rep.loc[output_rep.ies == "PUCV", var] = output_rep[var].div(1000)
        output_rep.loc[output_rep.ies == "UA", var] = output_rep[var].div(1000)
        output_rep.loc[output_rep.ies == "UTALCA", var] = output_rep[var].div(1000)

    # UOHiggins escribio en pesos
    variables = ["sub_item_c_2_1", "sub_item_c_2_2", "sub_item_c_3_1", "sub_item_c_3_2", "sub_item_c_3_3",
                 "sub_item_c_3_4"]
    for var in variables:
        output_rep.loc[output_rep.ies == "UDO", var] = output_rep[var].div(1000)

    # Valores negativos a positivo
    variables = ["sub_item_c_2_1", "sub_item_c_2_2", "sub_item_c_3_1", "sub_item_c_3_2", "sub_item_c_3_3",
                 "sub_item_c_3_4", "item_m_2", "item_m_3", "item_m_4", "sub_item_1_1_1", "sub_item_1_1_2",
                 "sub_item_1_1_3", "item_1_2", "item_2_1", "item_2_2", "item_2_3", "item_2_4", "item_2_5", "item_3_1",
                 "item_3_2", "item_3_3", "item_3_4", "item_3_5", "item_3_6", "item_4_1", "item_4_2"]
    for var in variables:
        output_rep[var] = output_rep[var].abs()

    # Porcentaje de actividades de pregrado y postgrado
    output_rep.loc[np.logical_and(output_rep.sub_item_c_1_1 == 0, output_rep.sub_item_c_1_2 == 0), "sub_item_c_1_2"] = 1

    # Arreglos
    # replace sub_item_c_1_1 = 1 if sub_item_c_1_1 > 1 & sub_item_c_1_1 != .
    output_rep.loc[
        np.logical_and(output_rep.sub_item_c_1_1 > 1, output_rep.sub_item_c_1_1.notna()), "sub_item_c_1_1"] = 1

    # replace sub_item_c_1_2 = 1 if sub_item_c_1_2 == .
    output_rep.loc[output_rep.sub_item_c_1_2.isna(), "sub_item_c_1_2"] = 1

    # replace sub_item_c_2_2 = 0 if sub_item_c_2_2 == .
    output_rep.loc[output_rep.sub_item_c_2_2.isna(), "sub_item_c_2_2"] = 0

    # Costo central per cápita
    output_rep["aux_cc"] = output_rep.sub_item_c_2_1 + output_rep.sub_item_c_3_1 + output_rep.sub_item_c_3_2 \
                           + output_rep.sub_item_c_3_3 + (output_rep.sub_item_c_2_2 * output_rep.sub_item_c_1_2) + \
                           (output_rep.sub_item_c_3_4 * output_rep.sub_item_c_1_2)

    # gen cc_pc = aux_cc / matricula_ies
    # output_rep["cc_pc"] = output_rep.aux_cc.div(output_rep.matricula_ies)

    # print(output_rep[["sub_item_c_2_1", "sub_item_c_3_1", "sub_item_c_3_2", "sub_item_c_3_3",
    #                   "sub_item_c_2_2", "sub_item_c_1_2", "sub_item_c_3_4", "sub_item_c_1_2"]].dtypes)

    output_rep.to_csv(os.path.join("data/", "output_rep.csv"), index=False, encoding="utf-8-sig", header=True)


def make_costos_centrales():
    path = "data/transform/costos_centrales/"
    files = os.listdir(path)

    dfs = []
    for file in files:
        ies = file.split(".")[0]
        df = pd.read_csv(os.path.join(path, file))
        df["ies"] = ies
        print(df)

        dfs.append(df)

    df = pd.concat(dfs, axis='rows')
    df.to_csv(os.path.join("data/transform/", "costos_centrales.csv"), index=False, encoding="utf-8-sig", header=True)


def make_costos_de_matricula():
    path = "data/transform/costos_de_matricula/"
    files = os.listdir(path)

    dfs = []
    for file in files:
        ies = file.split(".")[0]
        df = pd.read_csv(os.path.join(path, file))
        df["ies"] = ies
        print(df)

        dfs.append(df)

    df = pd.concat(dfs, axis='rows')
    df.to_csv(os.path.join("data/transform/", "costos_de_matricula.csv"), index=False, encoding="utf-8-sig",
              header=True)


def make_codigo_unico():
    path = "data/transform/codigo_unico/"
    files = os.listdir(path)

    dfs = []
    for file in files:
        ies = file.split(".")[0]
        df = pd.read_csv(os.path.join(path, file))
        df["ies"] = ies
        print(ies)

        dfs.append(df)

    df = pd.concat(dfs, axis='rows')
    df.to_csv(os.path.join("data/transform/", "codigo_unico.csv"), index=False, encoding="utf-8-sig", header=True)


def make_oferta_2020():
    oferta_2020 = get_dta_from_bbdd("oferta_2020")
    file = "oferta_2020.csv"
    oferta_2020.to_csv(os.path.join("data/BBDD/", file), index=False, encoding="utf-8-sig", header=True)


def replicate():
    """ replica para todas las ies a partir de la info entregada por las ies """
    extract_ies_all()
    transform_ies_to_df_all()
    make_codigo_unico()
    make_costos_centrales()
    make_costos_de_matricula()
    make_output()
    make_output_rep()


def replicate_ies(ies):
    """ replica solo para una ies (ej. UDV) """
    extract_ies(ies + ".xlsx")
    transform_ies_to_df(ies)
    make_codigo_unico()
    make_costos_centrales()
    make_costos_de_matricula()
    make_output()
    make_output_rep()


class Carrera:
    def __init__(self, carrera, year):
        self.carrera = carrera
        self.year = year

    def get_data(self):
        df = get_output_rep()
        df.anio = df.anio.astype(int)
        df = pd.DataFrame(df[df.anio == self.year])
        df = pd.DataFrame(df[df.carrera == self.carrera])
        return df

    def get_year(self):
        return self.year

    def get_matricula(self):
        df = self.get_data()
        df = df.groupby("institucion")["totalmatriculados"].sum()
        df = pd.DataFrame(df)
        df = df[df.totalmatriculados != 0]
        df["Matricula"] = df["totalmatriculados"]
        df = df[["Matricula"]]
        df = df.astype(int)
        return df

    def get_codigo_unico(self):
        df = self.get_data()
        df = pd.crosstab(index=df.institucion, columns="codigo_unico")
        return df

    def get_costos_directos_docencia(self):
        df = self.get_data()
        df["Planta (M$)"] = df["sub_item_1_1_1"]
        df["Contrata (M$)"] = df["sub_item_1_1_2"]
        df["Honorarios (M$)"] = df["sub_item_1_1_3"]
        df["Apoyo (M$)"] = df["item_1_2"]
        df["Costos Directos (M$)"] = df["sub_item_1_1_1"] + df["sub_item_1_1_2"] + df["sub_item_1_1_3"] + df["item_1_2"]

        df = df.groupby("institucion")["Planta (M$)", "Contrata (M$)", "Honorarios (M$)",
                                       "Apoyo (M$)", "Costos Directos (M$)"].sum()
        df = pd.DataFrame(df)
        df = df.astype(int)
        df = df.round(0)
        df = df[df["Costos Directos (M$)"] != 0]
        return df

    def get_costos_directos_docencia_por_matricula(self):
        costos_docencia = self.get_costos_directos_docencia()
        matricula = self.get_matricula()
        df = pd.merge(costos_docencia, matricula, how='left', left_index=True, right_index=True)

        df["Planta (M$)"] = df["Planta (M$)"].div(df["Matricula"])
        df["Contrata (M$)"] = df["Contrata (M$)"].div(df["Matricula"])
        df["Honorarios (M$)"] = df["Honorarios (M$)"].div(df["Matricula"])
        df["Apoyo (M$)"] = df["Apoyo (M$)"].div(df["Matricula"])
        df["Costos Directos (M$)"] = df["Costos Directos (M$)"].div(df["Matricula"])
        df = df.round(2)

        df = df.sort_values(by=['Costos Directos (M$)'], ascending=False)
        df = df[["Planta (M$)", "Contrata (M$)", "Honorarios (M$)", "Apoyo (M$)", "Costos Directos (M$)"]]
        return df

    def get_otros_costos_docencia(self):
        df = self.get_data()
        df["Apoyo docencia (M$)"] = df["item_2_1"] + df["item_2_3"]
        df["Mobiliario (M$)"] = df["item_2_2"] + df["item_2_4"]
        df["Equipos (M$)"] = df["item_2_5"]
        df["Otros Costos (M$)"] = df["item_2_1"] + df["item_2_2"] + df["item_2_3"] + df["item_2_4"] + df["item_2_5"]

        df = df.groupby("institucion")["Apoyo docencia (M$)",
                                       "Mobiliario (M$)", "Equipos (M$)",
                                       "Otros Costos (M$)"].sum()

        df = pd.DataFrame(df)
        df = df.astype(int)
        df = df.round(0)
        df = df[df["Otros Costos (M$)"] != 0]
        return df

    def get_otros_costos_docencia_por_matricula(self):
        otros_costos_docencia = self.get_otros_costos_docencia()
        matricula = self.get_matricula()
        df = pd.merge(otros_costos_docencia, matricula, how='left', left_index=True, right_index=True)

        df["Apoyo docencia (M$)"] = df["Apoyo docencia (M$)"].div(df["Matricula"])
        df["Mobiliario (M$)"] = df["Mobiliario (M$)"].div(df["Matricula"])
        df["Equipos (M$)"] = df["Equipos (M$)"].div(df["Matricula"])
        df["Otros Costos (M$)"] = df["Otros Costos (M$)"].div(df["Matricula"])
        df = df.round(2)

        df = df.sort_values(by=['Otros Costos (M$)'], ascending=False)
        df = df[["Apoyo docencia (M$)", "Mobiliario (M$)", "Equipos (M$)", "Otros Costos (M$)"]]
        return df

    def get_costos_indirectos(self):
        df = self.get_data()
        df["Pers. NO académico (M$)"] = df["item_3_1"]
        df["Docencia y Difusión (M$)"] = df["item_3_2"] + df["item_3_3"] + df["item_3_4"] + df["item_3_5"]
        df["Mantenimiento (M$)"] = df["item_3_6"]
        df["Costos Indirectos (M$)"] = df["item_3_1"] + df["item_3_2"] + df["item_3_3"] + df["item_3_4"] + df[
            "item_3_5"] + df["item_3_6"]

        df = df.groupby("institucion")["Pers. NO académico (M$)",
                                       "Docencia y Difusión (M$)",
                                       "Mantenimiento (M$)",
                                       "Costos Indirectos (M$)"].sum()

        df = pd.DataFrame(df)
        df = df.astype(int)
        df = df.round(0)
        df = df[df["Costos Indirectos (M$)"] != 0]
        return df

    def get_costos_indirectos_por_matricula(self):
        costos_indirectos = self.get_costos_indirectos()
        matricula = self.get_matricula()
        df = pd.merge(costos_indirectos, matricula, how='left', left_index=True, right_index=True)

        df["Pers. NO académico (M$)"] = df["Pers. NO académico (M$)"].div(df["Matricula"])
        df["Docencia y Difusión (M$)"] = df["Docencia y Difusión (M$)"].div(df["Matricula"])
        df["Mantenimiento (M$)"] = df["Mantenimiento (M$)"].div(df["Matricula"])
        df["Costos Indirectos (M$)"] = df["Costos Indirectos (M$)"].div(df["Matricula"])
        df = df.round(2)
        df = df.sort_values(by=['Costos Indirectos (M$)'], ascending=False)

        cols = len(df.columns) - 1
        col_name = "Costos Indirectos (M$)"
        last_col = df.pop(col_name)
        df.insert(cols, col_name, last_col)

        # elimina matricula
        df.pop("Matricula")
        return df


def get_output():
    """ este es un input para producir output_rep """
    df = pd.read_stata("data/output.dta")
    return df


def get_output_sub(save=False):
    """ toma output.dta y guarda output_sub.csv y retorna un df (output_sub.dta = output.dta)"""
    df = pd.read_stata("data/output.dta")
    if save:
        df.to_csv("data/output_sub.csv", index=False, encoding="utf-8", header=True)
    return df


def get_output_rep(save=False):
    """ toma output.dta y guarda output_rep.csv y retorna un df con los cambios  requeridos por la comision """
    df = pd.read_stata("data/output.dta")

    # cambio los 0 por NAN
    df = df.replace({'0': np.nan, 0: np.nan})

    if save:
        df.to_csv("data/output.csv", index=False, encoding="utf-8", header=True)
    return df


def main():
    df = pd.read_csv("reporte-costos/20200706_Matrícula_Ed_Superior_2020_PUBL_MRUN.csv")

    print(df.head())


if __name__ == "__main__":
    main()
